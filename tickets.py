from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from pydantic import BaseModel
from database import get_db, pwd_context
from auth import get_current_user
from ticket_logger import log_ticket_change
from typing import Optional, List
import os
import uuid

class TicketCreate(BaseModel):
    title: str
    description: str
    priority: str = "normal"
    type: str = "Прочие"

class TicketUpdate(BaseModel):
    status: Optional[str] = None
    deadline: Optional[str] = None
    priority: Optional[str] = None
    report: Optional[str] = None
    assigned_to: Optional[int] = None

class TicketExecutorInfo(BaseModel):
    """Информация об исполнителе на заявке"""
    id: int
    username: str
    full_name: str
    assigned_at: str
    assigned_by: Optional[int] = None

class TicketHistoryEntry(BaseModel):
    """Запись истории изменений заявки"""
    id: int
    change_type: str
    field_name: Optional[str] = None
    old_value: Optional[str] = None
    new_value: Optional[str] = None
    description: Optional[str] = None
    changed_by: int
    created_at: str

class ExecutorAssignRequest(BaseModel):
    """Для админа: добавление исполнителя"""
    user_id: int

class SettlementCreate(BaseModel):
    name: str

class StreetCreate(BaseModel):
    settlement_id: int
    name: str

class HouseCreate(BaseModel):
    street_id: int
    number: str

class ApartmentCreate(BaseModel):
    house_id: int
    number: str
    phone: Optional[str] = None

class UserCreate(BaseModel):
    username: str
    password: str
    full_name: str
    role: str = "executor"

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None

router = APIRouter(prefix="/tickets", tags=["tickets"])

@router.get("/")
async def get_tickets(request: Request):
    get_current_user(request)
    conn = get_db()
    tickets = conn.execute(
        "SELECT * FROM tickets ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return [dict(t) for t in tickets]

@router.post("/")
async def create_ticket(data: TicketCreate, request: Request):
    user = get_current_user(request)
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute(
        "INSERT INTO tickets (title, description, priority, type, created_by) VALUES (?, ?, ?, ?, ?)",
        (data.title, data.description, data.priority, data.type, user["id"])
    )
    conn.commit()

    # Получить ID созданной заявки
    ticket_id = cursor.lastrowid

    # Логировать создание заявки
    log_ticket_change(
        conn, ticket_id, "ticket_created", user["id"],
        description="Ticket created"
    )

    conn.close()
    return {"message": "OK", "ticket_id": ticket_id}

@router.get("/addresses")
async def get_addresses(request: Request):
    get_current_user(request)
    conn = get_db()
    addresses = conn.execute(
        "SELECT * FROM addresses ORDER BY settlement, street"
    ).fetchall()
    conn.close()
    return [dict(a) for a in addresses]

@router.get("/settlements")
async def get_settlements(request: Request):
    get_current_user(request)
    conn = get_db()
    settlements = conn.execute(
        "SELECT id, name FROM settlements ORDER BY name"
    ).fetchall()
    conn.close()
    return [dict(s) for s in settlements]

@router.get("/streets/{settlement_id}")
async def get_streets(settlement_id: int, request: Request):
    get_current_user(request)
    conn = get_db()
    streets = conn.execute(
        "SELECT id, name FROM streets WHERE settlement_id = ? ORDER BY name",
        (settlement_id,)
    ).fetchall()
    conn.close()
    return [dict(s) for s in streets]

@router.get("/houses/{street_id}")
async def get_houses(street_id: int, request: Request):
    get_current_user(request)
    conn = get_db()
    houses = conn.execute(
        "SELECT id, number FROM houses WHERE street_id = ? ORDER BY number",
        (street_id,)
    ).fetchall()
    conn.close()
    return [dict(h) for h in houses]

@router.post("/add-settlement")
async def add_settlement(data: SettlementCreate, request: Request):
    get_current_user(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO settlements (name) VALUES (?)",
            (data.name,)
        )
        conn.commit()
        return {"id": cursor.lastrowid, "name": data.name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@router.post("/add-street")
async def add_street(data: StreetCreate, request: Request):
    get_current_user(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO streets (settlement_id, name) VALUES (?, ?)",
            (data.settlement_id, data.name)
        )
        conn.commit()
        return {"id": cursor.lastrowid, "settlement_id": data.settlement_id, "name": data.name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@router.post("/add-house")
async def add_house(data: HouseCreate, request: Request):
    get_current_user(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO houses (street_id, number) VALUES (?, ?)",
            (data.street_id, data.number)
        )
        conn.commit()
        return {"id": cursor.lastrowid, "street_id": data.street_id, "number": data.number}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()

@router.get("/{ticket_id}")
async def get_ticket(ticket_id: int, request: Request):
    get_current_user(request)
    conn = get_db()
    ticket = conn.execute(
        "SELECT * FROM tickets WHERE id = ?", (ticket_id,)
    ).fetchone()

    if not ticket:
        conn.close()
        raise HTTPException(status_code=404, detail="Ticket not found")

    # Получить исполнителей
    executors = conn.execute("""
        SELECT u.id, u.username, u.full_name, te.assigned_at, te.assigned_by
        FROM ticket_executors te
        JOIN users u ON te.user_id = u.id
        WHERE te.ticket_id = ?
        ORDER BY te.assigned_at DESC
    """, (ticket_id,)).fetchall()

    # Получить историю изменений
    history = conn.execute("""
        SELECT id, change_type, field_name, old_value, new_value, description, changed_by, created_at
        FROM ticket_history
        WHERE ticket_id = ?
        ORDER BY created_at DESC
    """, (ticket_id,)).fetchall()

    conn.close()

    return {
        "ticket": dict(ticket),
        "executors": [dict(e) for e in executors],
        "history": [dict(h) for h in history]
    }

@router.put("/{ticket_id}")
async def update_ticket(ticket_id: int, data: TicketUpdate, request: Request):
    user = get_current_user(request)
    conn = get_db()

    # Получить текущее состояние заявки
    current_ticket = conn.execute(
        "SELECT * FROM tickets WHERE id = ?", (ticket_id,)
    ).fetchone()

    if not current_ticket:
        conn.close()
        raise HTTPException(status_code=404, detail="Ticket not found")

    current_ticket_dict = dict(current_ticket)

    # Проверки при переводе в статус "in_progress"
    if data.status == "in_progress" and current_ticket_dict.get("status") != "in_progress":
        executor_count = conn.execute(
            "SELECT COUNT(*) as count FROM ticket_executors WHERE ticket_id = ?",
            (ticket_id,)
        ).fetchone()["count"]

        if executor_count == 0:
            conn.close()
            raise HTTPException(
                status_code=400,
                detail="Cannot transition to in_progress without executors"
            )

    # Обновляем и логируем изменения
    if data.status is not None and data.status != current_ticket_dict.get("status"):
        conn.execute(
            "UPDATE tickets SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (data.status, ticket_id)
        )
        log_ticket_change(
            conn, ticket_id, "status_change", user["id"],
            field_name="status",
            old_value=current_ticket_dict.get("status"),
            new_value=data.status
        )

    if data.priority is not None and data.priority != current_ticket_dict.get("priority"):
        conn.execute(
            "UPDATE tickets SET priority=? WHERE id=?",
            (data.priority, ticket_id)
        )
        log_ticket_change(
            conn, ticket_id, "priority_update", user["id"],
            field_name="priority",
            old_value=current_ticket_dict.get("priority"),
            new_value=data.priority
        )

    if data.deadline is not None and data.deadline != current_ticket_dict.get("deadline"):
        conn.execute(
            "UPDATE tickets SET deadline=? WHERE id=?",
            (data.deadline, ticket_id)
        )
        log_ticket_change(
            conn, ticket_id, "deadline_update", user["id"],
            field_name="deadline",
            old_value=current_ticket_dict.get("deadline"),
            new_value=data.deadline
        )

    if data.report is not None and data.report != current_ticket_dict.get("report"):
        conn.execute(
            "UPDATE tickets SET report=? WHERE id=?",
            (data.report, ticket_id)
        )
        log_ticket_change(
            conn, ticket_id, "report_added", user["id"],
            field_name="report",
            new_value=data.report
        )

    if data.assigned_to is not None and data.assigned_to != current_ticket_dict.get("assigned_to"):
        # Update assigned_to in tickets table
        conn.execute(
            "UPDATE tickets SET assigned_to=? WHERE id=?",
            (data.assigned_to if data.assigned_to > 0 else None, ticket_id)
        )
        # Add to ticket_executors if not already there
        if data.assigned_to and data.assigned_to > 0:
            conn.execute(
                "INSERT OR IGNORE INTO ticket_executors (ticket_id, user_id, assigned_by) VALUES (?, ?, ?)",
                (ticket_id, data.assigned_to, user["id"])
            )
        log_ticket_change(
            conn, ticket_id, "assignment", user["id"],
            field_name="executor",
            old_value=str(current_ticket_dict.get("assigned_to")) if current_ticket_dict.get("assigned_to") else None,
            new_value=str(data.assigned_to) if data.assigned_to else None
        )

    conn.commit()
    conn.close()
    return {"message": "OK"}

@router.get("/{ticket_id}/comments")
async def get_comments(ticket_id: int, request: Request):
    get_current_user(request)
    conn = get_db()
    comments = conn.execute("""
        SELECT c.*, u.full_name FROM comments c
        JOIN users u ON c.user_id = u.id
        WHERE c.ticket_id = ?
        ORDER BY c.created_at ASC
    """, (ticket_id,)).fetchall()
    conn.close()
    result = []
    for c in comments:
        d = dict(c)
        d['photos'] = d['photo_path'].split(',') if d['photo_path'] else []
        result.append(d)
    return result

@router.post("/{ticket_id}/comments")
async def add_comment(
    ticket_id: int,
    request: Request,
    text: str = Form(...),
    photos: List[UploadFile] = File(default=[])
):
    user = get_current_user(request)
    photo_paths = []
    for photo in photos:
        if photo.filename:
            ext = photo.filename.split('.')[-1]
            filename = f"{uuid.uuid4()}.{ext}"
            filepath = f"uploads/{filename}"
            with open(filepath, 'wb') as f:
                content = await photo.read()
                f.write(content)
            photo_paths.append(filename)

    conn = get_db()
    conn.execute(
        "INSERT INTO comments (ticket_id, user_id, text, photo_path) VALUES (?, ?, ?, ?)",
        (ticket_id, user["id"], text, ','.join(photo_paths) if photo_paths else None)
    )
    conn.commit()
    conn.close()
    return {"message": "OK"}


# ───────────── EXECUTORS MANAGEMENT ─────────────

@router.post("/{ticket_id}/take")
async def take_ticket(ticket_id: int, request: Request):
    """Исполнитель берет заявку на себя"""
    user = get_current_user(request)
    conn = get_db()

    # Проверить существование заявки
    ticket = conn.execute(
        "SELECT * FROM tickets WHERE id = ?", (ticket_id,)
    ).fetchone()

    if not ticket:
        conn.close()
        raise HTTPException(status_code=404, detail="Ticket not found")

    ticket_dict = dict(ticket)

    try:
        # Добавить исполнителя
        conn.execute(
            "INSERT OR IGNORE INTO ticket_executors (ticket_id, user_id) VALUES (?, ?)",
            (ticket_id, user["id"])
        )

        # Если статус "new", переводим в "in_progress"
        status_changed = False
        if ticket_dict.get("status") == "new":
            conn.execute(
                "UPDATE tickets SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                ("in_progress", ticket_id)
            )
            log_ticket_change(
                conn, ticket_id, "status_change", user["id"],
                field_name="status",
                old_value="new",
                new_value="in_progress",
                description="Executor took the ticket"
            )
            status_changed = True

        # Логировать добавление исполнителя
        log_ticket_change(
            conn, ticket_id, "assignment", user["id"],
            field_name="executor",
            new_value=str(user["id"]),
            description="Self-assigned"
        )

        conn.commit()
        conn.close()

        return {
            "message": "OK",
            "status_changed": status_changed,
            "new_status": "in_progress" if status_changed else ticket_dict.get("status")
        }
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{ticket_id}/executors")
async def get_executors(ticket_id: int, request: Request):
    """Получить список исполнителей на заявке"""
    get_current_user(request)
    conn = get_db()

    executors = conn.execute("""
        SELECT u.id, u.username, u.full_name, te.assigned_at, te.assigned_by
        FROM ticket_executors te
        JOIN users u ON te.user_id = u.id
        WHERE te.ticket_id = ?
        ORDER BY te.assigned_at DESC
    """, (ticket_id,)).fetchall()

    conn.close()
    return [dict(e) for e in executors]


@router.get("/{ticket_id}/history")
async def get_ticket_history(ticket_id: int, request: Request, limit: int = 100):
    """Получить историю изменений заявки"""
    get_current_user(request)
    conn = get_db()

    history = conn.execute("""
        SELECT id, change_type, field_name, old_value, new_value, description, changed_by, created_at
        FROM ticket_history
        WHERE ticket_id = ?
        ORDER BY created_at DESC
        LIMIT ?
    """, (ticket_id, limit)).fetchall()

    conn.close()
    return [dict(h) for h in history]



# ───────────── ADMIN ENDPOINTS ─────────────
admin_router = APIRouter(prefix="/admin", tags=["admin"])

def check_admin(request: Request):
    """Проверка прав администратора"""
    user = get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    return user


# ПОЛЬЗОВАТЕЛИ
@admin_router.get("/users")
async def admin_get_users(request: Request):
    check_admin(request)
    conn = get_db()
    users = conn.execute(
        "SELECT id, username, full_name, role, created_at FROM users ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return [dict(u) for u in users]


@admin_router.get("/users/{user_id}")
async def admin_get_user(user_id: int, request: Request):
    check_admin(request)
    conn = get_db()
    user = conn.execute(
        "SELECT id, username, full_name, role, created_at FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()
    conn.close()
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    return dict(user)


@admin_router.post("/users")
async def admin_create_user(data: UserCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        hashed_password = pwd_context.hash(data.password)
        cursor = conn.execute(
            "INSERT INTO users (username, password, full_name, role) VALUES (?, ?, ?, ?)",
            (data.username, hashed_password, data.full_name, data.role)
        )
        conn.commit()
        return {"id": cursor.lastrowid, "username": data.username}
    except Exception as e:
        raise HTTPException(status_code=400, detail="Ошибка создания пользователя")
    finally:
        conn.close()


@admin_router.put("/users/{user_id}")
async def admin_update_user(user_id: int, data: UserUpdate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        if data.full_name:
            conn.execute("UPDATE users SET full_name = ? WHERE id = ?", (data.full_name, user_id))
        if data.role:
            conn.execute("UPDATE users SET role = ? WHERE id = ?", (data.role, user_id))
        if data.password:
            hashed = pwd_context.hash(data.password)
            conn.execute("UPDATE users SET password = ? WHERE id = ?", (hashed, user_id))
        conn.commit()
        return {"message": "OK"}
    except Exception as e:
        raise HTTPException(status_code=400, detail="Ошибка обновления")
    finally:
        conn.close()


@admin_router.delete("/users/{user_id}")
async def admin_delete_user(user_id: int, request: Request):
    check_admin(request)
    current_user = get_current_user(request)
    if current_user["id"] == user_id:
        raise HTTPException(status_code=400, detail="Нельзя удалить свой аккаунт")
    
    conn = get_db()
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return {"message": "OK"}


# ЗАЯВКИ
@admin_router.delete("/tickets/{ticket_id}")
async def admin_delete_ticket(ticket_id: int, request: Request):
    check_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM comments WHERE ticket_id = ?", (ticket_id,))
    conn.execute("DELETE FROM tickets WHERE id = ?", (ticket_id,))
    conn.commit()
    conn.close()
    return {"message": "OK"}


# ПОСЕЛЕНИЯ
@admin_router.post("/settlements")
async def admin_create_settlement(data: SettlementCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO settlements (name) VALUES (?)",
            (data.name,)
        )
        conn.commit()
        return {"id": cursor.lastrowid, "name": data.name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.put("/settlements/{settlement_id}")
async def admin_update_settlement(settlement_id: int, data: SettlementCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        conn.execute("UPDATE settlements SET name = ? WHERE id = ?", (data.name, settlement_id))
        conn.commit()
        return {"message": "OK"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.delete("/settlements/{settlement_id}")
async def admin_delete_settlement(settlement_id: int, request: Request):
    check_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM streets WHERE settlement_id = ?", (settlement_id,))
    conn.execute("DELETE FROM settlements WHERE id = ?", (settlement_id,))
    conn.commit()
    conn.close()
    return {"message": "OK"}


# УЛИЦЫ
@admin_router.get("/streets")
async def admin_get_streets(request: Request):
    check_admin(request)
    conn = get_db()
    streets = conn.execute("""
        SELECT s.id, s.settlement_id, s.name, se.name as settlement_name 
        FROM streets s 
        JOIN settlements se ON s.settlement_id = se.id 
        ORDER BY se.name, s.name
    """).fetchall()
    conn.close()
    return [dict(s) for s in streets]


@admin_router.post("/streets")
async def admin_create_street(data: StreetCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO streets (settlement_id, name) VALUES (?, ?)",
            (data.settlement_id, data.name)
        )
        conn.commit()
        return {"id": cursor.lastrowid}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.put("/streets/{street_id}")
async def admin_update_street(street_id: int, data: StreetCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        conn.execute(
            "UPDATE streets SET settlement_id = ?, name = ? WHERE id = ?",
            (data.settlement_id, data.name, street_id)
        )
        conn.commit()
        return {"message": "OK"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.delete("/streets/{street_id}")
async def admin_delete_street(street_id: int, request: Request):
    check_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM houses WHERE street_id = ?", (street_id,))
    conn.execute("DELETE FROM streets WHERE id = ?", (street_id,))
    conn.commit()
    conn.close()
    return {"message": "OK"}


# ДОМА
# ДОМА
@admin_router.get("/houses")
async def admin_get_houses(request: Request):
    check_admin(request)
    conn = get_db()
    houses = conn.execute("""
        SELECT h.id, h.street_id, h.number, s.name as street_name, se.name as settlement_name
        FROM houses h
        JOIN streets s ON h.street_id = s.id
        JOIN settlements se ON s.settlement_id = se.id
        ORDER BY se.name, s.name, h.number
    """).fetchall()
    conn.close()
    return [dict(h) for h in houses]


@admin_router.post("/houses")
async def admin_create_house(data: HouseCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO houses (street_id, number) VALUES (?, ?)",
            (data.street_id, data.number)
        )
        conn.commit()
        return {"id": cursor.lastrowid}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.put("/houses/{house_id}")
async def admin_update_house(house_id: int, data: HouseCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        conn.execute(
            "UPDATE houses SET street_id = ?, number = ? WHERE id = ?",
            (data.street_id, data.number, house_id)
        )
        conn.commit()
        return {"message": "OK"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.delete("/houses/{house_id}")
async def admin_delete_house(house_id: int, request: Request):
    check_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM apartments WHERE house_id = ?", (house_id,))
    conn.execute("DELETE FROM houses WHERE id = ?", (house_id,))
    conn.commit()
    conn.close()
    return {"message": "OK"}


# КВАРТИРЫ
@admin_router.get("/apartments")
async def admin_get_apartments(request: Request):
    check_admin(request)
    conn = get_db()
    apartments = conn.execute("""
        SELECT a.id, a.house_id, a.number, a.phone,
               se.name as settlement_name, s.name as street_name, h.number as house_number,
               (se.name || ' - ' || s.name || ' - ' || h.number) as full_address
        FROM apartments a
        JOIN houses h ON a.house_id = h.id
        JOIN streets s ON h.street_id = s.id
        JOIN settlements se ON s.settlement_id = se.id
        ORDER BY se.name, s.name, h.number, a.number
    """).fetchall()
    conn.close()
    return [dict(a) for a in apartments]


@admin_router.post("/apartments")
async def admin_create_apartment(data: ApartmentCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        cursor = conn.execute(
            "INSERT INTO apartments (house_id, number, phone) VALUES (?, ?, ?)",
            (data.house_id, data.number, data.phone)
        )
        conn.commit()
        return {"id": cursor.lastrowid}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.put("/apartments/{apartment_id}")
async def admin_update_apartment(apartment_id: int, data: ApartmentCreate, request: Request):
    check_admin(request)
    conn = get_db()
    try:
        conn.execute(
            "UPDATE apartments SET house_id = ?, number = ?, phone = ? WHERE id = ?",
            (data.house_id, data.number, data.phone, apartment_id)
        )
        conn.commit()
        return {"message": "OK"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        conn.close()


@admin_router.delete("/apartments/{apartment_id}")
async def admin_delete_apartment(apartment_id: int, request: Request):
    check_admin(request)
    conn = get_db()
    conn.execute("DELETE FROM apartments WHERE id = ?", (apartment_id,))
    conn.commit()
    conn.close()
    return {"message": "OK"}


# ИМПОРТ/ЭКСПОРТ БД (Excel)
SHEET_LABELS = {
    'users':       'Пользователи',
    'tickets':     'Заявки',
    'settlements': 'Поселения',
    'streets':     'Улицы',
    'houses':      'Дома',
    'apartments':  'Квартиры',
    'comments':    'Комментарии',
    'messages':    'Сообщения',
    'ticket_executors': 'Исполнители заявок',
    'ticket_history': 'История заявок',
}
SHEET_TABLES = {v: k for k, v in SHEET_LABELS.items()}

HEADER_BG    = '1a1a2e'
HEADER_FG    = 'FFFFFF'
ALT_ROW      = 'F0F2F5'
TABLE_ORDER  = ['users', 'tickets', 'settlements', 'streets', 'houses', 'apartments', 'comments', 'messages', 'ticket_executors', 'ticket_history']


def _build_excel(db_data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for table_name in TABLE_ORDER:
        table_data = db_data.get('tables', {}).get(table_name)
        if not table_data:
            continue
        cols = table_data['columns']
        rows = table_data['rows']
        label = SHEET_LABELS.get(table_name, table_name)
        ws = wb.create_sheet(title=label)

        # Заголовки
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = Font(bold=True, color=HEADER_FG, name='Arial', size=11)
            cell.fill = PatternFill('solid', start_color=HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[1].height = 22

        # Колонки с датами — форматируем как текст чтобы Excel не менял формат
        date_cols = {col for col in cols if any(kw in col for kw in ('_at', 'deadline', 'date'))}

        # Строки данных
        for ri, row in enumerate(rows, 2):
            bg = ALT_ROW if ri % 2 == 0 else 'FFFFFF'
            for ci, col in enumerate(cols, 1):
                val = row.get(col)
                if col in date_cols and val is not None and val != '':
                    val = str(val)
                cell = ws.cell(row=ri, column=ci, value=('' if val is None else val))
                cell.font = Font(name='Arial', size=10)
                cell.fill = PatternFill('solid', start_color=bg)
                cell.alignment = Alignment(vertical='center')
                cell.border = border
                if col in date_cols:
                    cell.number_format = '@'

        # Автоширина колонок
        for ci, col in enumerate(cols, 1):
            max_len = len(col)
            for row in rows:
                val = str(row.get(col) or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 55)

        ws.freeze_panes = 'A2'
        ws.sheet_properties.tabColor = '667eea'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_excel(content: bytes) -> dict:
    from openpyxl import load_workbook
    from io import BytesIO
    import datetime as dt

    def _coerce(val):
        """Привести значение из Excel к нужному типу для SQLite."""
        if val is None:
            return None
        # datetime/date → строка ISO
        if isinstance(val, (dt.datetime, dt.date)):
            return val.isoformat()
        # float без дробной части → int (Excel хранит id как 1.0)
        if isinstance(val, float):
            return int(val) if val == int(val) else val
        # bool → int
        if isinstance(val, bool):
            return int(val)
        return val

    wb = load_workbook(BytesIO(content), data_only=True)
    result = {'tables': {}}

    for sheet_name in wb.sheetnames:
        table_name = SHEET_TABLES.get(sheet_name)
        if not table_name:
            continue
        ws = wb[sheet_name]
        rows_iter = list(ws.iter_rows(values_only=True))
        if not rows_iter:
            continue
        columns = [str(c) for c in rows_iter[0]]
        rows = []
        for row in rows_iter[1:]:
            if all(v is None for v in row):
                continue
            rows.append({columns[i]: _coerce(row[i]) for i in range(len(columns))})
        result['tables'][table_name] = {'columns': columns, 'rows': rows}

    return result


@admin_router.get("/export-excel")
async def export_excel(request: Request):
    from database import export_db
    from fastapi.responses import Response
    from datetime import datetime
    check_admin(request)
    data = export_db()
    excel_bytes = _build_excel(data)
    filename = f"dispatch_db_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@admin_router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), request: Request = None):
    from database import import_db
    check_admin(request)
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Файл должен быть в формате .xlsx")
    try:
        content = await file.read()
        data = _parse_excel(content)
        if not data['tables']:
            raise HTTPException(status_code=400, detail="Не найдено ни одной известной таблицы в файле")
        import_db(data)
        imported = list(data['tables'].keys())
        return {"message": f"Импортировано таблиц: {len(imported)}", "tables": imported}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка импорта: {str(e)}")


# ───────────── EXECUTORS MANAGEMENT ENDPOINTS (ADMIN) ─────────────

@admin_router.post("/tickets/{ticket_id}/executors")
async def admin_add_executor(ticket_id: int, data: ExecutorAssignRequest, request: Request):
    """Админ добавляет исполнителя на заявку"""
    user = check_admin(request)
    conn = get_db()

    try:
        # Проверить существование заявки
        ticket = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
        if not ticket:
            conn.close()
            raise HTTPException(status_code=404, detail="Ticket not found")

        # Проверить существование пользователя
        executor = conn.execute("SELECT * FROM users WHERE id = ?", (data.user_id,)).fetchone()
        if not executor:
            conn.close()
            raise HTTPException(status_code=404, detail="User not found")

        # Добавить исполнителя
        conn.execute(
            "INSERT OR IGNORE INTO ticket_executors (ticket_id, user_id, assigned_by) VALUES (?, ?, ?)",
            (ticket_id, data.user_id, user["id"])
        )

        # Логировать добавление
        log_ticket_change(
            conn, ticket_id, "assignment", user["id"],
            field_name="executor",
            new_value=str(data.user_id),
            description=f"Admin assigned executor {executor['full_name']}"
        )

        conn.commit()
        conn.close()
        return {"message": "OK"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))


@admin_router.delete("/tickets/{ticket_id}/executors/{user_id}")
async def admin_remove_executor(ticket_id: int, user_id: int, request: Request):
    """Админ удаляет исполнителя с заявки"""
    user = check_admin(request)
    conn = get_db()

    try:
        # Проверить существование
        executor_record = conn.execute(
            "SELECT * FROM ticket_executors WHERE ticket_id = ? AND user_id = ?",
            (ticket_id, user_id)
        ).fetchone()

        if not executor_record:
            conn.close()
            raise HTTPException(status_code=404, detail="Executor assignment not found")

        # Удалить исполнителя
        conn.execute(
            "DELETE FROM ticket_executors WHERE ticket_id = ? AND user_id = ?",
            (ticket_id, user_id)
        )

        # Логировать удаление
        log_ticket_change(
            conn, ticket_id, "unassignment", user["id"],
            field_name="executor",
            old_value=str(user_id),
            description="Admin removed executor"
        )

        # Проверить, остались ли исполнители
        remaining_executors = conn.execute(
            "SELECT COUNT(*) as count FROM ticket_executors WHERE ticket_id = ?",
            (ticket_id,)
        ).fetchone()["count"]

        # Если нет исполнителей и статус был "in_progress", переводим в "new"
        if remaining_executors == 0:
            ticket = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
            if ticket and dict(ticket).get("status") == "in_progress":
                conn.execute(
                    "UPDATE tickets SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    ("new", ticket_id)
                )
                log_ticket_change(
                    conn, ticket_id, "status_change", user["id"],
                    field_name="status",
                    old_value="in_progress",
                    new_value="new",
                    description="Status reverted to new (no executors)"
                )

        conn.commit()
        conn.close()
        return {"message": "OK"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))